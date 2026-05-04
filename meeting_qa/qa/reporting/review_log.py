"""review_log.py — Single-tab xlsx workbook and JSON trace.

Workbook columns (one row per claim adjudication):
  agenda_item | claim_text | claim_type | blocking_candidate |
  source_excerpt | judge1_verdict | judge1_rationale |
  judge2_verdict | judge2_rationale | reviewer_notes

JSON trace: machine-readable record of the full routing decision.
"""
from __future__ import annotations

import json
from pathlib import Path

from qa.engine.models import ClaimAdjudication, DeterministicResult, ProjectInput, RouteDecision


# ── Workbook ──────────────────────────────────────────────────────────────────

_COLUMNS = [
    "agenda_item",
    "claim_text",
    "claim_type",
    "blocking_candidate",
    "source_excerpt",
    "judge1_verdict",
    "judge1_rationale",
    "judge2_verdict",
    "judge2_rationale",
    "reviewer_notes",
]


def write_workbook(
    adjudications: list[ClaimAdjudication],
    project_input: ProjectInput,
    path: Path,
    det_results: list[DeterministicResult] | None = None,
) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print("  [review_log] openpyxl not installed — workbook skipped")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Claim Review"

    # Header row
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    for col, header in enumerate(_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    wrap = Alignment(wrap_text=True, vertical="top")
    for adj in adjudications:
        claim = adj.claim
        row = [
            claim.item_title,
            claim.claim_text,
            claim.claim_type,
            "Yes" if claim.blocking_candidate else "No",
            adj.bundle.matched_local_span[:400] if adj.bundle.matched_local_span else "",
            adj.phase1.accuracy_category if adj.phase1 else "",
            adj.phase1.rationale if adj.phase1 else "",
            adj.phase2.accuracy_category if adj.phase2 else "",
            adj.phase2.rationale if adj.phase2 else "",
            "",  # reviewer_notes — blank for human
        ]
        ws.append(row)
        for col in range(1, len(_COLUMNS) + 1):
            ws.cell(row=ws.max_row, column=col).alignment = wrap

    # Column widths
    _widths = [30, 50, 22, 12, 50, 28, 50, 28, 50, 20]
    for col, width in enumerate(_widths, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col)
        ].width = width

    # Colour rows by outcome
    from openpyxl.styles import PatternFill
    red_fill   = PatternFill("solid", fgColor="FFCCCC")
    amber_fill = PatternFill("solid", fgColor="FFF2CC")

    for row_idx, adj in enumerate(adjudications, 2):
        verdict = adj.final_verdict
        if verdict in ("Incorrect", "Not in Source — Unresolved"):
            fill = red_fill
        elif verdict in ("Unverifiable", "Not in Source — Verified Elsewhere", "Extrapolating"):
            fill = amber_fill
        else:
            fill = None
        if fill:
            for col in range(1, len(_COLUMNS) + 1):
                ws.cell(row=row_idx, column=col).fill = fill

    # ── Det Checks sheet ──────────────────────────────────────────────────────
    if det_results:
        import json as _json
        ws2 = wb.create_sheet("Det Checks")
        det_headers = [
            "check_name", "severity", "reason",
            "llm_verdict", "llm_rationale", "details",
        ]
        for col, h in enumerate(det_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        red_fill2   = PatternFill("solid", fgColor="FFCCCC")
        amber_fill2 = PatternFill("solid", fgColor="FFF2CC")
        green_fill2 = PatternFill("solid", fgColor="CCFFCC")

        for r in det_results:
            verdict = r.llm_verdict or ("-" if not r.needs_llm_verification else "Not verified")
            row2 = [
                r.check_name,
                "Blocking" if r.blocks else "Annotation",
                r.reason,
                verdict,
                r.llm_rationale or "",
                _json.dumps(r.details) if r.details else "",
            ]
            ws2.append(row2)
            row_idx2 = ws2.max_row
            for col in range(1, len(det_headers) + 1):
                ws2.cell(row=row_idx2, column=col).alignment = wrap
            if r.blocks:
                if r.llm_verdict == "Cleared":
                    fill2 = green_fill2
                elif r.llm_verdict == "Confirmed" or not r.needs_llm_verification:
                    fill2 = red_fill2
                else:
                    fill2 = amber_fill2
                for col in range(1, len(det_headers) + 1):
                    ws2.cell(row=row_idx2, column=col).fill = fill2

        det_widths = [28, 12, 60, 18, 60, 40]
        for col, w in enumerate(det_widths, 1):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    wb.save(path)
    print(f"  [review_log] Workbook: {path}")


# ── JSON trace ────────────────────────────────────────────────────────────────

def write_trace(
    route: RouteDecision,
    project_input: ProjectInput,
    path: Path,
) -> None:
    """Write a compact machine-readable trace of the routing decision."""
    trace = {
        "document_id": project_input.document_id,
        "document_type": project_input.document_type,
        "route": {
            "final_status": route.final_status,
            "reason_code": route.reason_code,
            "human_reason": route.human_reason,
            "triggered_by": route.triggered_by,
        },
        "deterministic": [
            {
                "check_name": r.check_name,
                "blocks": r.blocks,
                "reason": r.reason,
                "needs_llm_verification": r.needs_llm_verification,
                "llm_verdict": r.llm_verdict,
                "llm_rationale": r.llm_rationale,
            }
            for r in route.deterministic_results
        ],
        "claims": [
            {
                "index": adj.claim.index,
                "item_slug": adj.claim.item_slug,
                "source_field": adj.claim.source_field,
                "claim_text": adj.claim.claim_text,
                "claim_type": adj.claim.claim_type,
                "weight_tier": adj.claim.weight_tier,
                "blocking_candidate": adj.claim.blocking_candidate,
                "should_skip": adj.claim.should_skip,
                "skip_reason": adj.claim.skip_reason,
                "phase1": (
                    {
                        "judge": adj.phase1.judge_name,
                        "verdict": adj.phase1.accuracy_category,
                        "rationale": adj.phase1.rationale,
                    }
                    if adj.phase1 else None
                ),
                "phase2": (
                    {
                        "judge": adj.phase2.judge_name,
                        "verdict": adj.phase2.accuracy_category,
                        "rationale": adj.phase2.rationale,
                    }
                    if adj.phase2 else None
                ),
                "final_verdict": adj.final_verdict,
                "blocks": (
                    adj.claim.blocking_candidate
                    and adj.phase2 is not None
                    and adj.final_verdict not in {"Accurate", "Directionally Consistent", "Extrapolating", "Modeled"}
                ),
            }
            for adj in route.all_adjudications
        ],
    }
    path.write_text(json.dumps(trace, indent=2), encoding="utf-8")
    print(f"  [review_log] Trace: {path}")
